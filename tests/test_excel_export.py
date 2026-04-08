"""Tests for Excel (.xlsx) export."""
from pathlib import Path

import pytest

from advisoryops.excel_export import export_excel, COLUMNS


@pytest.fixture
def sample_issues():
    return [
        {
            "issue_id": "CVE-2024-1234",
            "priority": "P0",
            "score": 145,
            "title": "Critical vulnerability in Device X",
            "summary": "A critical RCE vulnerability affects Device X firmware.",
            "cves": ["CVE-2024-1234"],
            "sources": ["cisa-icsma", "nvd-cve-api"],
            "vendor": "Acme Medical",
            "severity": "critical",
            "handling_warnings": ["do not reboot without vendor guidance"],
            "evidence_gaps": ["affected versions unclear"],
            "evidence_completeness": 0.75,
            "generated_by": "hybrid",
        },
        {
            "issue_id": "CVE-2024-5678",
            "priority": "P2",
            "score": 55,
            "title": "Medium info disclosure in System Y",
            "summary": "Information disclosure via unauthenticated endpoint.",
            "cves": ["CVE-2024-5678"],
            "sources": ["certcc-vulnotes"],
            "vendor": "HealthCorp",
            "severity": "medium",
            "handling_warnings": [],
            "evidence_gaps": [],
            "evidence_completeness": 0.5,
            "generated_by": "deterministic",
        },
    ]


def test_export_excel_creates_file(sample_issues, tmp_path):
    out = tmp_path / "test.xlsx"
    result = export_excel(sample_issues, out)
    assert result.exists()
    assert result.suffix == ".xlsx"


def test_export_excel_returns_path(sample_issues, tmp_path):
    out = tmp_path / "test.xlsx"
    result = export_excel(sample_issues, out)
    assert result == out


def test_export_excel_valid_xlsx(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.title == "Issues"


def test_export_excel_column_count(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.max_column == len(COLUMNS)


def test_export_excel_header_names(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS) + 1)]
    assert headers == COLUMNS


def test_export_excel_row_count(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    # 1 header + 2 data rows
    assert ws.max_row == 3


def test_export_excel_priority_values(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "P0"
    assert ws.cell(row=3, column=2).value == "P2"


def test_export_excel_issue_id_values(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "CVE-2024-1234"


def test_export_excel_empty_issues(tmp_path):
    out = tmp_path / "empty.xlsx"
    result = export_excel([], out)
    assert result.exists()
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.max_row == 1  # header only


def test_export_excel_frozen_panes(sample_issues, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "test.xlsx"
    export_excel(sample_issues, out)
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.freeze_panes == "A2"


def test_export_excel_sanitizes_illegal_characters(tmp_path):
    """Control characters \\x00-\\x08, \\x0B, \\x0C, \\x0E-\\x1F should not crash openpyxl."""
    from openpyxl import load_workbook
    issues = [
        {
            "issue_id": "CVE-2024-CTRL",
            "priority": "P2",
            "score": 50,
            "title": "Device with \x00null \x07bell \x0Bvtab chars",
            "summary": "Images may be missing\x00 when a system parameter\x07 is set.\x1F End.",
            "cves": ["CVE-2024-CTRL"],
            "sources": ["openfda-recalls-historical"],
            "vendor": "Test\x08Corp",
        },
    ]
    out = tmp_path / "ctrl.xlsx"
    # This should NOT raise IllegalCharacterError
    result = export_excel(issues, out)
    assert result.exists()
    wb = load_workbook(str(out))
    ws = wb.active
    # Verify data is present but control chars are stripped
    title_val = ws.cell(row=2, column=4).value
    assert "null" in title_val
    assert "bell" in title_val
    assert "\x00" not in title_val
    assert "\x07" not in title_val
